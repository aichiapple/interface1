import time

from openpyxl import Workbook,load_workbook

from common.text_util import read_text_handel
from pathlib import Path

from common.yaml_util import *

base_dir = Path(__file__).parent.parent

class ExcelUtil(object):
    def __init__(self,excel_path):
        self.wb=load_workbook(excel_path)
        self.template="""{"id":0,"url":"","case_name":"","header":"","method":"","body":"",
        "expect":"","actual":"","valiadate":""},"""

    def read_excel(self):
        """读取excel，处理数据，返回一个格式处理后的字典"""
        value=[]
        for sheetname in self.wb.sheetnames:
        #获取默认的工作表
            ws=self.wb[sheetname]
            default_sheet_name=ws.title
            case_num=len(list(ws.values))-1
            case_list=list(ws.values)
            #去除标题行
            case_list.pop(0)
            cases_template=self.template * case_num
            print(type(cases_template))

            cases_template_list=eval("[" + cases_template[:-1] + "]")
            #此时cases_template_list为list，元素为模板，个数为列表的元素个数

            for i in range(len(case_list)):
                cases_template_list[i]["id"]=case_list[i][0]
                cases_template_list[i]["url"] = case_list[i][1]
                cases_template_list[i]["case_name"] = case_list[i][2]
                cases_template_list[i]["header"] = case_list[i][3]
                cases_template_list[i]["method"] = case_list[i][4]
                cases_template_list[i]["body"] = case_list[i][5]
                cases_template_list[i]["expect"] = case_list[i][6]
                cases_template_list[i]["actual"] = case_list[i][7]
                cases_template_list[i]["valiadate"] = case_list[i][8]
            value.append({default_sheet_name : cases_template_list})
        print(value)
        return value

    def write_excel(self,resultWriteSheet):
        """结果写入excel"""
        l_reponse,l_ispass=read_text_handel()

        i=0
        j=0
        #for sheetname in self.wb.sheetnames:
        ws=self.wb[resultWriteSheet]
        #实际结果列
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=8,min_col=8):
            for cell in row:
                cell.value=l_reponse[i]
                i+=1
        #是否通过列处理
        for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=9,min_col=9):
            for cell in row:
                cell.value=l_ispass[j]
        save_path="%s/output/run_result_excel/运行结果_%s.xlsx"(self.base_dir,time.strftime("%Y%m%d_%H:%M:%S"))
        self.wb.save(save_path)
    def handle_excel(self):
        yaml_data=self.read_excel()
        for item in yaml_data:
            sheetnames=item.keys()
            for sheetname in sheetnames:
                write_yaml("%s/data/data_yaml/%s.yaml" % (base_dir, sheetname), item[sheetname])


        # for yaml_fileName,value in enumerate(yaml_data):
        #     write_yaml("%s/data/data_yaml/%s.yaml" % (base_dir,yaml_fileName),value)

if __name__ == '__main__':
    ExcelUtil("%s/data/case_excel/MCC自动化测试表格.xlsx" % base_dir).handle_excel()

